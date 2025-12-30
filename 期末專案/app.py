from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import jwt
from functools import wraps
from datetime import datetime, timedelta
import os
import io
import csv
import openpyxl # ✅ 新增：用於產生 Excel
# ✅ 新增：匯入模型以供 Excel 下載使用
from models import SessionLocal, Parcel, Customer, Account, TrackingEvent

from db_operations import (
    initialize_database,
    append_customer,
    read_customers,
    update_customer,
    append_parcel,
    read_parcels,
    update_parcel_amount,
    update_parcel_status,
    append_tracking_event,
    read_tracking_events,
    append_account,
    read_accounts,
    find_account,
    read_all_events_for_search,
    delete_parcel_by_tracking,
)

app = Flask(__name__)
CORS(app, resources={r"/*": {"origins": "*"}}, allow_headers=["Content-Type", "Authorization"])

SECRET_KEY = "my_secret_key_for_jwt_12345"

# ------------------------------------------------
# 初始化預設帳號
# ------------------------------------------------
def init_default_accounts():
    accounts = read_accounts()
    defaults = {
        "staff1": {"password": "staff123", "role": "staff"},
        "admin1": {"password": "admin123", "role": "admin"},
        "driver1": {"password": "driver123", "role": "driver"},
        "warehouse1": {"password": "warehouse123", "role": "warehouse"},  
        "test1": {"password": "test123", "role": "customer"},
    }
    for username, info in defaults.items():
        if username not in accounts:
            append_account({
                "username": username,
                "password": info["password"],
                "role": info["role"]
            })

# ------------------------------------------------
# JWT 驗證裝飾器
# ------------------------------------------------
def token_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = None
        if "Authorization" in request.headers:
            auth = request.headers["Authorization"]
            if auth.startswith("Bearer "):
                token = auth.split(" ")[1]

        if not token:
            return jsonify({"error": "缺少 JWT Token"}), 401

        try:
            data = jwt.decode(token, SECRET_KEY, algorithms=["HS256"])
            request.user = data
        except jwt.ExpiredSignatureError:
            return jsonify({"error": "Token 已過期"}), 401
        except Exception:
            return jsonify({"error": "無效 Token"}), 401

        return f(*args, **kwargs)

    return decorated

# ------------------------------------------------
# 登入
# ------------------------------------------------
@app.route("/api/auth/login", methods=["POST"])
def login():
    data = request.get_json() or {}
    username = data.get("username")
    password = data.get("password")

    if not username or not password:
        return jsonify({"error": "請提供帳號與密碼"}), 400

    account = find_account(username)
    if not account:
        return jsonify({"error": "帳號不存在"}), 401

    if account["password"] != password:
        return jsonify({"error": "密碼錯誤"}), 401

    token = jwt.encode(
        {
            "username": username,
            "role": account["role"],
            "exp": datetime.utcnow() + timedelta(hours=4),
        },
        SECRET_KEY,
        algorithm="HS256",
    )

    return jsonify({
        "message": "登入成功",
        "username": username,
        "role": account["role"],
        "token": token,
    })

# ------------------------------------------------
# 客戶註冊
# ------------------------------------------------
@app.route("/api/auth/register", methods=["POST"])
def register_customer_account():
    data = request.get_json() or {}
    username = data.get("username")
    password = data.get("password")
    name = data.get("name") or ""
    phone = data.get("phone") or ""
    email = data.get("email") or ""
    address = data.get("address") or ""
    customer_type = data.get("customer_type") or "NON_CONTRACT"  
    billing_preference = data.get("billing_preference") or "COD"  

    if not username or not password:
        return jsonify({"error": "請提供帳號與密碼"}), 400

    if find_account(username):
        return jsonify({"error": "帳號已存在"}), 400

    append_account({
        "username": username,
        "password": password,
        "role": "customer"
    })
    append_customer({
        "account": username,
        "name": name,
        "phone": phone,
        "email": email,
        "address": address,
        "customer_type": customer_type,
        "billing_preference": billing_preference,
    })

    return jsonify({"message": "註冊成功", "username": username})

# ------------------------------------------------
# 客戶資料 API
# ------------------------------------------------
@app.route("/api/customers", methods=["POST"])
@token_required
def create_customer():
    if request.user.get("role") not in ["staff", "admin"]:
        return jsonify({"error": "權限不足"}), 403
    data = request.get_json() or {}
    append_customer(data)
    return jsonify({"message": "客戶已建立", "customer_id": data.get("account")})

@app.route("/api/customers", methods=["GET"])
@token_required
def list_customers():
    role = request.user.get("role")
    if role not in ["staff", "admin"]:
        return jsonify({"error": "權限不足"}), 403
    customers = read_customers() or []
    return jsonify(customers)

@app.route("/api/customers/<account>", methods=["PUT"])
@token_required
def edit_customer(account):
    if request.user.get("role") not in ["admin", "staff"]:
        return jsonify({"error": "只有管理員與作業人員可以修改客戶資料"}), 403
    data = request.get_json() or {}
    update_customer(account, data)
    return jsonify({"message": "客戶資料已更新"})

# ------------------------------------------------
# 建立包裹 (含重量體積後端檢查)
# ------------------------------------------------
@app.route("/api/parcels", methods=["POST"])
@token_required
def create_parcel():
    data = request.get_json() or {}

    sender_id = (data.get("sender") or data.get("sender_id") or request.user.get("username"))
    recipient_name = (data.get("receiver") or data.get("receiver_name") or data.get("recipient_name"))
    recipient_address = (data.get("receiverAddress") or data.get("receiver_address") or data.get("recipient_address") or "")
    
    weight = data.get("weight")
    volume = data.get("volume")
    
    package_type = data.get("package_type") or "中型箱" 
    declared_value = data.get("declared_value") or 0    
    contents = data.get("contents") or "一般貨物"        
    service_type = data.get("service_type") or "標準速遞"

    # 後端防呆：檢查重量
    try:
        w_val = float(weight)
        if w_val <= 0:
            return jsonify({"error": "重量必須大於 0"}), 400
    except (TypeError, ValueError):
        return jsonify({"error": "重量格式錯誤"}), 400

    # 後端防呆：檢查體積 (若是純數字字串)
    if volume:
        try:
            v_val = float(volume)
            if v_val < 0:
                return jsonify({"error": "體積不能為負數"}), 400
        except ValueError:
            pass

    if not sender_id or not recipient_name:
        return jsonify({"error": "缺少寄件人或收件人"}), 400

    today = datetime.now().strftime("%Y%m%d")
    rand4 = str(datetime.now().microsecond)[-4:]
    tracking_number = f"TRK-{today}-{rand4}"
    
    created_at = datetime.now()  # 直接傳入 datetime 物件

    record = {
        "tracking_number": tracking_number,
        "sender_id": sender_id,
        "recipient_name": recipient_name,
        "recipient_address": recipient_address,
        "weight": weight,
        "package_type": package_type,      
        "declared_value": declared_value,  
        "contents": contents,              
        "service_type": service_type,
        "status": "建立包裹",
        "amount": None,
        "created_at": created_at,
    }

    append_parcel(record)
    
    # 記錄事件
    event_id = f"EVT-{datetime.now().strftime('%Y%m%d%H%M%S%f')}"
    append_tracking_event({
        "event_id": event_id,
        "tracking_number": tracking_number,
        "event_type": "建立包裹",
        "timestamp": created_at,  # ✅ 修改：直接傳入 datetime
        "location": "系統",
        "operator": request.user.get("username"),
        "description": f"包裹由 {sender_id} 建立"
    })

    return jsonify({
        "message": "包裹建立成功",
        "tracking_no": tracking_number,
        "package": {
            **record,
            "created_at": created_at.strftime("%Y-%m-%d %H:%M:%S")  # 回傳給前端時格式化
        },
    }), 201

# ------------------------------------------------
# 更新包裹金額
# ------------------------------------------------

@app.route("/api/parcels/amount", methods=["POST"])
@token_required
def set_parcel_amount():
    data = request.get_json() or {}
    tracking = data.get("tracking_number") or data.get("tracking_no")
    amount = data.get("amount")
    pay_method = data.get("payment_method", "cash")
    new_service_type = data.get("service_type") # 取得前端傳來的服務類型

    if not tracking:
        return jsonify({"error": "缺少追蹤編號"}), 400
    if amount is None:
        return jsonify({"error": "缺少金額"}), 400

    try:
        amount_val = float(amount)
        if amount_val < 0:
             return jsonify({"error": "金額不能為負數"}), 400
    except (TypeError, ValueError):
        return jsonify({"error": "金額格式錯誤"}), 400

    # 決定付款狀態文字
    payment_status_text = "已付款(線上)"
    if pay_method in ["cash", "cod"]:
        payment_status_text = "待付款(貨到付款)"
    elif pay_method == "monthly":
        payment_status_text = "月結帳單"
    elif pay_method == "prepaid":
        payment_status_text = "已付款(預付)"

    # 直接使用 SessionLocal 更新資料庫
    db = SessionLocal()
    try:
        parcel = db.query(Parcel).filter(Parcel.tracking_number == tracking).first()
        if not parcel:
            return jsonify({"error": "找不到包裹"}), 404
            
        # ✅ 重點：同時更新金額、付款狀態、服務類型
        parcel.amount = amount_val
        parcel.payment_status = payment_status_text 
        
        if new_service_type:
            parcel.service_type = new_service_type
            
        db.commit()
        
        # 記錄事件
        event_id = f"EVT-{datetime.now().strftime('%Y%m%d%H%M%S%f')}"
        desc = f"金額: {amount_val}, 方式: {pay_method}, 狀態: {payment_status_text}"
        if new_service_type:
            desc += f", 服務類型更新為: {new_service_type}"

        new_event = TrackingEvent(
            event_id=event_id,
            tracking_number=tracking,
            event_type="計費完成",
            timestamp=datetime.now(),
            location="系統",
            operator=request.user.get("username"),
            description=desc
        )
        db.add(new_event)
        db.commit()
        
    except Exception as e:
        db.rollback()
        print(f"Error updating amount: {e}")
        return jsonify({"error": "更新失敗"}), 500
    finally:
        db.close()

    return jsonify({
        "message": "金額與狀態已更新",
        "tracking_number": tracking,
        "amount": amount_val,
        "payment_status": payment_status_text,
        "service_type": new_service_type
    })

# ------------------------------------------------
# 更新包裹狀態 (含權限與異常鎖定)
# ------------------------------------------------
@app.route("/api/parcels/status", methods=["POST"])
@token_required
def set_parcel_status():
    role = request.user.get("role")
    if role == "customer":
        return jsonify({"error": "客戶無權修改包裹狀態"}), 403

    data = request.get_json() or {}
    tracking = data.get("tracking_number") or data.get("tracking_no")
    new_status = (data.get("status") or "").strip()
    
    location = data.get("location", "")
    vehicle_id = data.get("vehicle_id", "")
    warehouse_id = data.get("warehouse_id", "")
    description = data.get("description", "")

    if not tracking or not new_status:
        return jsonify({"error": "資料不全"}), 400

    parcels = read_parcels()
    current_parcel = next((p for p in parcels if p["tracking_number"] == tracking), None)
    if not current_parcel:
        return jsonify({"error": "找不到該追蹤編號"}), 404

    current_status = current_parcel.get("status", "")
    ABNORMAL_STATUSES = ["遺失", "損毀", "退回"]
    
    if current_status in ABNORMAL_STATUSES and new_status not in ["處理中", "退回"]:
        if role != "admin":
             return jsonify({"error": f"包裹處於 '{current_status}' 狀態,無法進行一般更新"}), 400

    if role == "driver":
        allowed = ["已裝車", "配送中", "已送達", "延誤", "遺失", "損毀"]
        if new_status not in allowed:
            return jsonify({"error": "司機無法執行此狀態變更"}), 403
            
    if role == "warehouse":
        allowed = ["已收件", "進入倉儲", "已裝車", "退回", "損毀"]
        if new_status not in allowed:
            return jsonify({"error": "倉儲人員無法執行此狀態變更"}), 403

    ok = update_parcel_status(tracking, new_status)
    if not ok:
        return jsonify({"error": "更新失敗"}), 404

    event_id = f"EVT-{datetime.now().strftime('%Y%m%d%H%M%S%f')}"
    append_tracking_event({
        "event_id": event_id,
        "tracking_number": tracking,
        "event_type": new_status,
        "timestamp": datetime.now(),  # ✅ 修改：直接傳入 datetime
        "location": location,
        "vehicle_id": vehicle_id,
        "warehouse_id": warehouse_id,
        "operator": request.user.get("username"),
        "description": description or f"狀態變更為: {new_status}"
    })

    return jsonify({"message": "狀態已更新", "status": new_status})

# ------------------------------------------------
# 刪除包裹
# ------------------------------------------------
@app.route("/api/parcels/<tracking_no>", methods=["DELETE"])
@token_required
def delete_parcel(tracking_no):
    role = request.user.get("role")
    if role not in ["admin", "staff"]:
        return jsonify({"error": "權限不足，無法刪除"}), 403

    success = delete_parcel_by_tracking(tracking_no)
    if success:
        return jsonify({"message": f"包裹 {tracking_no} 已刪除"})
    else:
        return jsonify({"error": "找不到該包裹"}), 404

# ------------------------------------------------
# 查詢與下載
# ------------------------------------------------
@app.route("/api/parcels/<tracking_no>/history", methods=["GET"])
@token_required
def get_parcel_history(tracking_no):
    events = read_tracking_events(tracking_no)
    if not events:
        return jsonify({"message": "查無追蹤紀錄", "events": []}), 200
    return jsonify({"tracking_number": tracking_no, "events": events})

@app.route("/records", methods=["GET"])
@token_required
def list_records():
    current_user = request.user
    role = current_user.get("role")
    username = current_user.get("username")
    vehicle_filter = request.args.get("vehicle_id")
    warehouse_filter = request.args.get("warehouse_id")

    parcels = read_parcels() or []

    if role == "customer":
        parcels = [p for p in parcels if p.get("sender_id") == username]
    
    allowed_tracking_nums = None
    if vehicle_filter or warehouse_filter:
        all_events = read_all_events_for_search()
        allowed_tracking_nums = set()
        for e in all_events:
            if vehicle_filter and vehicle_filter.lower() in str(e.get("vehicle_id") or "").lower():
                allowed_tracking_nums.add(e["tracking_number"])
            if warehouse_filter and warehouse_filter.lower() in str(e.get("warehouse_id") or "").lower():
                allowed_tracking_nums.add(e["tracking_number"])
    
    rows = []
    for p in parcels:
        if allowed_tracking_nums is not None:
            if p.get("tracking_number") not in allowed_tracking_nums:
                continue
        created_at = p.get("created_at") or ""
        date_only = created_at.split(" ")[0] if created_at else ""
        rows.append({
            "tracking_no": p.get("tracking_number"),
            "sender_id": p.get("sender_id"),
            "receiver_name": p.get("recipient_name"),
            "recipient_address": p.get("recipient_address"), # ✅ 確保前端能拿到地址
            "weight": p.get("weight"),
            "package_type": p.get("package_type", ""),
            "date": date_only,
            "amount": p.get("amount"),
            "status": p.get("status"),
        })
    return jsonify(rows)

# ------------------------------------------------
# ✅ 終極版 Excel 下載：日誌 + 包裹 + 客戶 + 帳號
# ------------------------------------------------
@app.route("/api/download", methods=["GET"])
@token_required
def download_excel():
    role = request.user.get("role")
    if role == "customer":
        return jsonify({"error": "權限不足"}), 403
    
    db = SessionLocal()
    try:
        wb = openpyxl.Workbook()
        
        # 分頁 1: 物流日誌 (TrackingEvent) - 這是您最想要的日誌
        ws_log = wb.active
        ws_log.title = "物流日誌(操作紀錄)"
        ws_log.append(['發生時間', '操作人員', '事件類型', '追蹤編號', '地點', '車輛/倉儲', '備註', '寄件人'])
        
        events = db.query(TrackingEvent).order_by(TrackingEvent.timestamp.desc()).all()
        for e in events:
            sender_id = e.parcel.sender_id if e.parcel else "(包裹已刪除)"
            ws_log.append([
                e.timestamp.strftime("%Y-%m-%d %H:%M:%S") if e.timestamp else "",
                e.operator or "系統自動",
                e.event_type,
                e.tracking_number,
                e.location,
                f"{e.vehicle_id or ''} {e.warehouse_id or ''}".strip(),
                e.description,
                sender_id
            ])

        # 分頁 2: 包裹清單 (Parcel)
        ws_parcel = wb.create_sheet("包裹清單")
        ws_parcel.append([
            '追蹤編號', '寄件人', '收件人', '收件地址', '重量', 
            '包裹類型', '申報價值', '內容物', '服務類型', 
            '狀態', '金額', '付款狀態', '建立時間'
        ])
        parcels = db.query(Parcel).all()
        for p in parcels:
            ws_parcel.append([
                p.tracking_number, p.sender_id, p.recipient_name, p.recipient_address,
                p.weight, p.package_type, p.declared_value, p.contents, p.service_type,
                p.status, p.amount, p.payment_status,
                p.created_at.strftime("%Y-%m-%d %H:%M:%S") if p.created_at else ""
            ])

        # 分頁 3: 客戶資料 (Customer)
        ws_cust = wb.create_sheet("客戶資料")
        ws_cust.append(['帳號', '姓名', '電話', 'Email', '地址', '客戶類型', '帳單偏好'])
        customers = db.query(Customer).all()
        for c in customers:
            ws_cust.append([
                c.account, c.name, c.phone, c.email, c.address, c.customer_type, c.billing_preference
            ])

        # 分頁 4: 系統帳號 (Account)
        ws_acc = wb.create_sheet("系統帳號")
        ws_acc.append(['帳號', '角色', '建立時間'])
        accounts = db.query(Account).all()
        for a in accounts:
            ws_acc.append([
                a.username, a.role, 
                a.created_at.strftime("%Y-%m-%d %H:%M:%S") if a.created_at else ""
            ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='logistics_full_data.xlsx'
        )
        
    except Exception as e:
        print(f"匯出失敗: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        db.close()

# ✅ 修改：使用資料庫初始化
initialize_database()
init_default_accounts()

if __name__ == "__main__":
    app.run(debug=True)